"""
utils/excel_importer.py
-------------------------
Imports Transfers (with their Tools, Part Numbers, and the most
commonly pre-filled Preparation fields) from an .xlsx workbook, so a
team already tracking transfers in a spreadsheet can bulk-load that
history into the app instead of re-entering it by hand.

Expected layout: one sheet, one row per (TRF Number, Tool Number, Part
Number) combination - the exact same shape produced by
`reports/excel_export.export_transfers_to_excel`, plus additional
optional columns for the raw editable fields of each Preparation
sub-module (see `IMPORT_TEMPLATE_COLUMNS` below). Only "TRF Number" and
"Tool Number" are required; every other column is optional and simply
left at its default when absent or blank. Header matching is
case-insensitive and ignores surrounding whitespace, so minor
formatting differences in the source sheet don't break the import.

Behaviour notes:
  * Rows are grouped by TRF Number (first-seen order). All transfer- and
    PTT/E2E-level columns are read from the FIRST row of each group.
  * Tool-level columns (Safety Stock, Training) are read from the first
    row in which that Tool Number appears within the group.
  * Every row contributes its own Part Number (if present) with its own
    Raw Material / Pre-check / Applicator / Counter Part columns.
  * Import is ADDITIVE: re-running it on the same file creates new
    Transfer records rather than updating existing ones with matching
    TRF Numbers, exactly like re-importing the same spreadsheet twice
    would in most simple tools. Duplicate TRF Numbers are allowed by
    the data model, so this is safe but worth knowing before re-running.
"""
from __future__ import annotations

from datetime import datetime, date
from typing import Optional

import openpyxl

from database.base import new_session
from models.transfer import Transfer, Tool, PartNumber
from services import transfer_service as svc
from services import progress_service
from utils import constants

# Column groups shown to the user when they download the import
# template (see reports/excel_export.export_import_template) and
# recognised on import. Each tuple is (header, applies_to) purely for
# documentation/template generation; the importer itself just looks
# headers up by name.
TRANSFER_COLUMNS = [
    "TRF Number", "Planned Transfer Date", "Actual Transfer Date", "Transfer Type",
    "Activity", "Sender Location", "Receiver Location", "Technology",
]
PTT_COLUMNS = [
    "PTT Internal Status", "PTT Responsible", "PTT Internal Due Date",
    "PTT Internal Approval Date", "PTT Comments",
]
E2E_COLUMNS = [
    "E2E Kickoff Week", "E2E Kickoff Status", "E2E PCN PPAP Week", "E2E PCN PPAP Status",
    "E2E PCN Decision", "E2E PCN Status", "E2E SOP Week", "E2E SOP Status", "E2E Comments",
]
TOOL_COLUMNS = [
    "Tool Number",
    "Safety Stock Required", "SS Start Week", "SS Number Of Weeks",
    "SS Required Quantity", "SS Built Quantity", "SS Finish Week",
    "Training Required", "Training Planned Week", "Training Duration",
    "Training Invitation Sent", "Training Status",
]
PART_NUMBER_COLUMNS = [
    "Part Number",
    "RM Subgroup", "RM Setup", "RM Order", "RM Availability", "RM Due Date", "RM Comment",
    "Pre-check PE Responsible", "Pre-check Samples Before", "Pre-check PE Requirement",
    "Pre-check Measurement Report", "Pre-check Feedback", "Pre-check Due Date",
    "Applicator PE", "Applicator Urgency", "Applicator Required", "Applicator",
    "Applicator Crimping Specification", "Applicator Terminal", "Applicator Wire Section",
    "Applicator Number Of Parts", "Applicator Required Approvals",
    "Applicator Available Location", "Applicator Crimping Request", "Applicator Comments",
    "Counter Part PE", "Counter Part", "Counter Part Terminal", "Counter Part Crimping",
    "Counter Part Terminal Request", "Counter Part Status", "Counter Part Comments",
]
IMPORT_TEMPLATE_COLUMNS = TRANSFER_COLUMNS + PTT_COLUMNS + E2E_COLUMNS + TOOL_COLUMNS + PART_NUMBER_COLUMNS

REQUIRED_COLUMNS = ["TRF Number", "Tool Number"]

_YES_VALUES = {"yes", "y", "true", "1"}


class ImportError_(Exception):
    """Raised for structural problems (missing required columns, etc.)."""


def _norm(text) -> str:
    return str(text).strip() if text is not None else ""


def _norm_header(text) -> str:
    return _norm(text).lower()


def _clean(value) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _to_date(value) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _to_bool(value, default: bool = False) -> bool:
    text = _clean(value).lower()
    if not text:
        return default
    return text in _YES_VALUES


def _to_float(value, default: float = 0.0) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def _to_int(value, default: int = 0) -> int:
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def _pick(value, options: list[str], default: str) -> str:
    """Case-insensitive match of a raw cell value against a fixed
    option list (e.g. status vocabularies); falls back to `default` if
    the cell is blank or doesn't match any option."""
    text = _clean(value)
    if not text:
        return default
    for opt in options:
        if opt.lower() == text.lower():
            return opt
    return default


class ExcelImporter:
    def __init__(self, xlsx_path: str, session=None):
        self.xlsx_path = xlsx_path
        self.session = session or new_session()
        self._owns_session = session is None
        self.wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    def run(self) -> dict:
        ws = self.wb.worksheets[0]
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [_norm(h) for h in header_row]
        header_lookup = {_norm_header(h): i for i, h in enumerate(headers) if h}

        missing = [c for c in REQUIRED_COLUMNS if _norm_header(c) not in header_lookup]
        if missing:
            raise ImportError_(
                f"The workbook is missing required column(s): {', '.join(missing)}. "
                f"Download the import template from Settings for the expected layout."
            )

        def cell(row_values, col_name):
            idx = header_lookup.get(_norm_header(col_name))
            return row_values[idx] if idx is not None and idx < len(row_values) else None

        # Group rows by TRF Number, preserving first-seen order.
        groups: dict[str, list] = {}
        order: list[str] = []
        skipped_rows = 0
        for row_values in ws.iter_rows(min_row=2, values_only=True):
            if row_values is None or all(v is None for v in row_values):
                continue
            trf = _clean(cell(row_values, "TRF Number"))
            tool_number = _clean(cell(row_values, "Tool Number"))
            if not trf or not tool_number:
                skipped_rows += 1
                continue
            groups.setdefault(trf, [])
            if trf not in order:
                order.append(trf)
            groups[trf].append(row_values)

        transfers_created = 0
        tools_created = 0
        part_numbers_created = 0

        for trf in order:
            rows = groups[trf]
            first = rows[0]

            transfer = Transfer(
                trf_number=trf,
                planned_transfer_date=_to_date(cell(first, "Planned Transfer Date")),
                actual_transfer_date=_to_date(cell(first, "Actual Transfer Date")),
                transfer_type=_pick(cell(first, "Transfer Type"), constants.TRANSFER_TYPES, constants.TRANSFER_TYPES[0]),
                activity=_pick(cell(first, "Activity"), constants.ACTIVITIES, constants.ACTIVITIES[0]),
                sender_location=_clean(cell(first, "Sender Location")),
                receiver_location=_clean(cell(first, "Receiver Location")),
                technology=_clean(cell(first, "Technology")),
            )
            svc.ensure_related_records(transfer)

            # PTT (transfer-level, from first row)
            ptt = transfer.ptt_approval
            ptt.internal_status = _pick(cell(first, "PTT Internal Status"), constants.STATUS_NOT_ONGOING_APPROVED, ptt.internal_status)
            ptt.internal_responsible = _clean(cell(first, "PTT Responsible")) or ptt.internal_responsible
            due = _to_date(cell(first, "PTT Internal Due Date"))
            if due:
                ptt.internal_due_date = due
            appr = _to_date(cell(first, "PTT Internal Approval Date"))
            if appr:
                ptt.internal_approval_date = appr
            ptt.internal_comments = _clean(cell(first, "PTT Comments")) or ptt.internal_comments

            # E2E (transfer-level, from first row)
            e2e = transfer.e2e_followup
            e2e.kickoff_week = _clean(cell(first, "E2E Kickoff Week")) or e2e.kickoff_week
            e2e.kickoff_status = _pick(cell(first, "E2E Kickoff Status"), constants.STATUS_CALL_GENERIC, e2e.kickoff_status)
            e2e.pcn_ppap_week = _clean(cell(first, "E2E PCN PPAP Week")) or e2e.pcn_ppap_week
            e2e.pcn_ppap_status = _pick(cell(first, "E2E PCN PPAP Status"), constants.STATUS_CALL_GENERIC, e2e.pcn_ppap_status)
            e2e.pcn_decision = _to_bool(cell(first, "E2E PCN Decision"), e2e.pcn_decision)
            e2e.pcn_status = _pick(cell(first, "E2E PCN Status"), constants.STATUS_NOT_SENT_SENT_ONGOING_APPROVED, e2e.pcn_status)
            e2e.sop_week = _clean(cell(first, "E2E SOP Week")) or e2e.sop_week
            e2e.sop_status = _pick(cell(first, "E2E SOP Status"), constants.STATUS_CALL_GENERIC, e2e.sop_status)
            e2e.comments = _clean(cell(first, "E2E Comments")) or e2e.comments

            # Tools + Part Numbers
            tools_by_number: dict[str, Tool] = {}
            for row_values in rows:
                tool_number = _clean(cell(row_values, "Tool Number"))
                if not tool_number:
                    continue
                is_new_tool = tool_number not in tools_by_number
                tool = tools_by_number.get(tool_number)
                if tool is None:
                    tool = Tool(tool_number=tool_number)
                    transfer.tools.append(tool)
                    tools_by_number[tool_number] = tool
                    tools_created += 1

                if is_new_tool:
                    from models.preparation import SafetyStock, Training
                    tool.safety_stock = SafetyStock(
                        required=_to_bool(cell(row_values, "Safety Stock Required")),
                        start_week=_clean(cell(row_values, "SS Start Week")),
                        number_of_weeks=_to_int(cell(row_values, "SS Number Of Weeks")),
                        required_quantity=_to_float(cell(row_values, "SS Required Quantity")),
                        built_quantity=_to_float(cell(row_values, "SS Built Quantity")),
                        finish_week=_clean(cell(row_values, "SS Finish Week")),
                    )
                    tool.training = Training(
                        required=_to_bool(cell(row_values, "Training Required")),
                        planned_week=_clean(cell(row_values, "Training Planned Week")),
                        duration=_clean(cell(row_values, "Training Duration")),
                        invitation_sent=_to_bool(cell(row_values, "Training Invitation Sent")),
                        status=_pick(cell(row_values, "Training Status"), constants.STATUS_NOT_ONGOING_DONE, "Not Started"),
                    )

                part_number = _clean(cell(row_values, "Part Number"))
                if not part_number:
                    continue

                pn = PartNumber(part_number=part_number)
                tool.part_numbers.append(pn)
                part_numbers_created += 1

                from models.preparation import RawMaterial, PreCheck, Applicator, CounterPart
                pn.raw_material = RawMaterial(
                    subgroup_status=_pick(cell(row_values, "RM Subgroup"), constants.STATUS_NA_ONGOING_DONE, "NA"),
                    setup_status=_pick(cell(row_values, "RM Setup"), constants.STATUS_NA_ONGOING_DONE, "NA"),
                    order_status=_pick(cell(row_values, "RM Order"), constants.STATUS_NA_ONGOING_DONE, "NA"),
                    availability_status=_pick(cell(row_values, "RM Availability"), constants.STATUS_NA_ONGOING_DONE, "NA"),
                    due_date=_to_date(cell(row_values, "RM Due Date")),
                    comment=_clean(cell(row_values, "RM Comment")),
                )
                pn.pre_check = PreCheck(
                    pe_responsible=_clean(cell(row_values, "Pre-check PE Responsible")),
                    samples_before_status=_pick(cell(row_values, "Pre-check Samples Before"), constants.STATUS_NA_ONGOING_RECEIVED, "NA"),
                    pe_requirement_status=_pick(cell(row_values, "Pre-check PE Requirement"), constants.STATUS_NA_ONGOING_RECEIVED, "NA"),
                    measurement_report_status=_pick(cell(row_values, "Pre-check Measurement Report"), constants.STATUS_NA_ONGOING_RECEIVED, "NA"),
                    feedback_status=_pick(cell(row_values, "Pre-check Feedback"), constants.STATUS_NA_ONGOING_ACCEPTED_REJECTED, "NA"),
                    due_date=_to_date(cell(row_values, "Pre-check Due Date")),
                )

                if transfer.activity == "Stamping":
                    pn.applicator = Applicator(
                        pe=_clean(cell(row_values, "Applicator PE")),
                        urgency=_pick(cell(row_values, "Applicator Urgency"), constants.URGENCY_LEVELS, "Medium"),
                        applicator_required=_to_bool(cell(row_values, "Applicator Required"), True),
                        applicator=_clean(cell(row_values, "Applicator")),
                        crimping_specification=_clean(cell(row_values, "Applicator Crimping Specification")),
                        terminal=_clean(cell(row_values, "Applicator Terminal")),
                        wire_section=_clean(cell(row_values, "Applicator Wire Section")),
                        number_of_parts=_to_int(cell(row_values, "Applicator Number Of Parts")),
                        required_approvals=_clean(cell(row_values, "Applicator Required Approvals")),
                        applicator_available_location=_clean(cell(row_values, "Applicator Available Location")),
                        crimping_request=_pick(cell(row_values, "Applicator Crimping Request"), constants.STATUS_NOT_ONGOING_DONE, "Not Started"),
                        comments=_clean(cell(row_values, "Applicator Comments")),
                    )
                else:
                    pn.counter_part = CounterPart(
                        pe=_clean(cell(row_values, "Counter Part PE")),
                        counter_part=_clean(cell(row_values, "Counter Part")),
                        terminal=_clean(cell(row_values, "Counter Part Terminal")),
                        crimping=_clean(cell(row_values, "Counter Part Crimping")),
                        terminal_request=_pick(cell(row_values, "Counter Part Terminal Request"), constants.STATUS_NOT_ONGOING_DONE, "Not Started"),
                        status_field=_pick(cell(row_values, "Counter Part Status"), constants.STATUS_NOT_ONGOING_DONE, "Not Started"),
                        comments=_clean(cell(row_values, "Counter Part Comments")),
                    )

            self.session.add(transfer)
            svc.ensure_related_records(transfer)
            self.session.flush()
            progress_service.refresh_transfer(self.session, transfer)
            svc.log_activity(self.session, transfer.id, "Imported", f"Imported from {self.xlsx_path}")
            transfers_created += 1

        self.session.commit()

        return {
            "transfers_created": transfers_created,
            "tools_created": tools_created,
            "part_numbers_created": part_numbers_created,
            "rows_skipped": skipped_rows,
        }

    def close(self):
        if self._owns_session:
            self.session.close()


def import_workbook(xlsx_path: str) -> dict:
    """Convenience entry point used by the Settings view."""
    importer = ExcelImporter(xlsx_path)
    try:
        return importer.run()
    finally:
        importer.close()
