"""
reports/excel_export.py
-------------------------
Exports Transfers to a formatted .xlsx workbook. Since a Transfer can
contain multiple Tools, each with multiple Part Numbers, the export is
flattened to one row per (Transfer, Tool, Part Number) combination -
the same shape as the original tracking spreadsheets - with a Transfer
that has no tools/PNs yet still getting a single summary row.
"""
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from models.transfer import Transfer
import config

HEADER_FILL = PatternFill("solid", fgColor="0F5FA8")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="DCE2EA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLUMNS = [
    "TRF Number", "Planned Transfer Date", "Transfer Type", "Activity",
    "Sender Location", "Receiver Location", "Technology", "Status",
    "Preparation Progress %", "Release Progress %",
    "Tool Number", "Part Number",
    "PTT Status", "Safety Stock Status", "RM Status", "Pre-check Status",
    "E2E Status", "Applicator/CP Status", "Training Status",
]


def _fmt_date(d):
    return d.isoformat() if d else ""


def export_import_template(file_path: str | None = None) -> str:
    """Generates a blank .xlsx with the exact headers the importer
    recognises (utils.excel_importer.IMPORT_TEMPLATE_COLUMNS), plus one
    filled-in example row, so users know exactly what to fill in before
    using Settings -> Import from Excel."""
    from utils.excel_importer import IMPORT_TEMPLATE_COLUMNS

    if file_path is None:
        file_path = str(Path(config.EXPORTS_DIR) / "transfer_import_template.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"

    ws.append(IMPORT_TEMPLATE_COLUMNS)
    for col_idx in range(1, len(IMPORT_TEMPLATE_COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = "A2"

    example = {
        "TRF Number": "TRF-26-00001",
        "Planned Transfer Date": "2026-09-15",
        "Transfer Type": "2-Step",
        "Activity": "Stamping",
        "Sender Location": "TE Woerth",
        "Receiver Location": "TE Tangier",
        "Technology": "MQS 1.0",
        "PTT Internal Status": "Ongoing",
        "Tool Number": "11-1058975",
        "Safety Stock Required": "Yes",
        "SS Required Quantity": "200",
        "SS Built Quantity": "50",
        "Part Number": "7-1452668-3",
        "RM Subgroup": "Done",
        "Pre-check Feedback": "Ongoing",
    }
    ws.append([example.get(h, "") for h in IMPORT_TEMPLATE_COLUMNS])

    widths = [16 if len(h) < 14 else 20 for h in IMPORT_TEMPLATE_COLUMNS]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    notes = wb.create_sheet("Notes")
    notes["A1"] = "Only 'TRF Number' and 'Tool Number' are required. Every other column is optional."
    notes["A2"] = "One row = one (TRF Number, Tool Number, Part Number) combination."
    notes["A3"] = "Transfer/PTT/E2E-level fields are read from the first row of each TRF Number."
    notes["A4"] = "Tool-level fields (Safety Stock, Training) are read from the first row of each Tool Number."
    notes["A5"] = "Applicator columns apply when Activity = Stamping; Counter Part columns when Activity = Molding."
    notes["A6"] = "Yes/No columns accept: Yes, No, True, False, 1, 0 (case-insensitive)."
    notes["A7"] = "Dates accept YYYY-MM-DD, DD/MM/YYYY, DD-MM-YYYY, or MM/DD/YYYY."
    notes.column_dimensions["A"].width = 100

    wb.save(file_path)
    return file_path


def export_transfers_to_excel(transfers: list[Transfer], file_path: str | None = None) -> str:
    if file_path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_path = str(Path(config.EXPORTS_DIR) / f"transfers_export_{ts}.xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Transfers"

    ws.append(COLUMNS)
    for col_idx in range(1, len(COLUMNS) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.freeze_panes = "A2"

    for t in transfers:
        ptt_status = t.ptt_approval.overall_status() if t.ptt_approval else ""
        e2e_status = "Done" if (t.e2e_followup and t.e2e_followup.progress_pct() >= 100) else ("Ongoing" if t.e2e_followup and t.e2e_followup.progress_pct() > 0 else "Not Started")

        rows_added = 0
        for tool in t.tools:
            ss_status = tool.safety_stock.status() if tool.safety_stock else ""
            tr_status = tool.training.status if tool.training else ""
            for pn in tool.part_numbers:
                rm_status = pn.raw_material.overall_status() if pn.raw_material else ""
                pc_status = pn.pre_check.feedback_status if pn.pre_check else ""
                if t.activity == "Stamping":
                    ap_status = pn.applicator.status() if pn.applicator else ""
                else:
                    ap_status = pn.counter_part.status_field if pn.counter_part else ""
                ws.append([
                    t.trf_number, _fmt_date(t.planned_transfer_date), t.transfer_type, t.activity,
                    t.sender_location, t.receiver_location, t.technology, t.status,
                    t.preparation_progress, t.release_progress,
                    tool.tool_number, pn.part_number,
                    ptt_status, ss_status, rm_status, pc_status, e2e_status, ap_status, tr_status,
                ])
                rows_added += 1
        if rows_added == 0:
            ws.append([
                t.trf_number, _fmt_date(t.planned_transfer_date), t.transfer_type, t.activity,
                t.sender_location, t.receiver_location, t.technology, t.status,
                t.preparation_progress, t.release_progress,
                "", "", ptt_status, "", "", "", e2e_status, "", "",
            ])

    for row_cells in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row_cells:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center")

    widths = [14, 16, 12, 11, 16, 16, 14, 12, 12, 12, 14, 14, 12, 14, 10, 12, 10, 14, 12]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(file_path)
    return file_path
